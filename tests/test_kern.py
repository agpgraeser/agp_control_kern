"""Rauchtests des Kerns (die ausführliche Abdeckung liegt historisch in den
Test-Suiten der Anwendungen, v. a. RegelkreisZustandsDemo/test_sim_core.py)."""

import numpy as np
import pytest

from agp_control_kern import (SimError, parse_matrix, parse_vector, parse_poly,
                              parse_complex_list, parse_zeitkonstanten,
                              zpk_to_tf, zk_to_tf, tf_to_ss,
                              matrix_zu_text, vektor_zu_text)


def test_parse_matrix_und_vector():
    A = parse_matrix("[0 1; -2 -3]")
    np.testing.assert_allclose(A, [[0, 1], [-2, -3]])
    b = parse_vector("0; 1", 2, "b")
    np.testing.assert_allclose(b, [0, 1])
    with pytest.raises(SimError):
        parse_matrix("[1 2; 3]")


def test_parse_poly_und_complex():
    np.testing.assert_allclose(parse_poly("[1 3 2]", "N"), [1, 3, 2])
    w = parse_complex_list("-1+2i; -1-2i", "Pole")
    assert w[0] == complex(-1, 2)


def test_zpk_und_zk_to_tf():
    z, n = zpk_to_tf([], [-1, -2], 2.0)
    np.testing.assert_allclose(z, [2.0])
    np.testing.assert_allclose(n, [1, 3, 2])
    z2, n2 = zk_to_tf("", "(s+1)(2s+1)", 5.0)
    np.testing.assert_allclose(z2, [5.0])
    np.testing.assert_allclose(n2, [2, 3, 1])


def test_zeitkonstanten_parser():
    np.testing.assert_allclose(
        parse_zeitkonstanten("(2*s+1)(0.5s^2 + 0.1s + 1)", "T"),
        [1.0, 0.7, 2.1, 1.0])


def test_tf_to_ss_rnf():
    erg = tf_to_ss([1.0], [1.0, 3.0, 2.0], "rnf")
    np.testing.assert_allclose(erg["A"], [[0, 1], [-2, -3]])
    assert erg["g0_tf"] == pytest.approx(0.5)
    assert erg["a_text"].startswith("[")


def test_texte():
    assert matrix_zu_text(np.array([[0.5, 1], [2, 3]])) == "[0.5 1; 2 3]"
    assert vektor_zu_text(np.array([1.0, 2.5]), "; ") == "1; 2.5"


# ── Projektdatei: Totzeit-Feld t_t im System-Blatt (Formatversion 1, additiv) ──

import io
from openpyxl import load_workbook
from agp_control_kern import projektdatei

_SYS_BASIS = {"systemname": "PT1", "a_text": "[-1]", "b_text": "1", "c_text": "1",
              "u_min": 0, "u_max": 1, "k_s": 1, "y_min": 0}


def test_projektdatei_t_t_roundtrip():
    """Totzeit t_t im System-Blatt übersteht Schreiben → Lesen."""
    p = projektdatei.neu(fallname="TotzeitTest")
    p["system"] = {**_SYS_BASIS, "t_t": 2.5}
    zurueck = projektdatei.lesen(projektdatei.schreiben(p))
    assert float(zurueck["system"]["t_t"]) == pytest.approx(2.5)


def test_projektdatei_altdatei_ohne_t_t_liest_ohne_fehler():
    """Altdatei ohne Totzeit-Zeile darf nicht scheitern (t_t ist optional)."""
    p = projektdatei.neu(fallname="Alt")
    p["system"] = dict(_SYS_BASIS)
    roh = projektdatei.schreiben(p)
    wb = load_workbook(io.BytesIO(roh))       # Totzeit-Zeile entfernen = Altdatei
    ws = wb["System"]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == "Totzeit T_t":
            ws.delete_rows(r)
            break
    buf = io.BytesIO()
    wb.save(buf)
    zurueck = projektdatei.lesen(buf.getvalue())   # darf NICHT werfen
    assert "t_t" not in zurueck["system"]
