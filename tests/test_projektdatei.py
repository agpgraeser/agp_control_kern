"""Tests der AGP-Projektdatei (Formatversion 1, Spec V1)."""

import io

import pytest
from openpyxl import Workbook, load_workbook

from agp_control_kern import projektdatei as pd
from agp_control_kern import ProjektdateiError


BEISPIEL_SYSTEM = {
    "systemname": "PT2-Test", "eingabeform": "ss",
    "a_text": "[0 1; -2 -3]", "b_text": "0; 1", "c_text": "1 0",
    "u_min": "0", "u_max": "100", "k_s": "0.8", "y_min": "20",
    "beschreibung": "Testfall",
}


def test_neu_schreiben_lesen_roundtrip():
    p = pd.neu("Fall-A", "Beschreibung", programm="Test")
    p["system"] = dict(BEISPIEL_SYSTEM)
    inhalt = pd.schreiben(p, programm="Test", aktion="System geschrieben")
    q = pd.lesen(inhalt)
    assert q["meta"]["formatversion"] == 1
    assert q["meta"]["fallname"] == "Fall-A"
    # Historie: "angelegt" + "System geschrieben"
    aktionen = [h["aktion"] for h in q["meta"]["historie"]]
    assert aktionen == ["Projekt angelegt", "System geschrieben"]
    assert q["system"]["a_text"] == "[0 1; -2 -3]"
    assert q["stoerungen"] is None and q["modelle"] is None


def test_alle_tabellen_roundtrip():
    p = pd.neu("Fall-B", programm="Test")
    p["system"] = dict(BEISPIEL_SYSTEM)
    p["stoerungen"] = [{"Nr": 1, "Typ": "sinus", "Angriff": "ausgang",
                        "Aktiv": "ja", "Amplitude": 0.5, "Startzeit": 10,
                        "Frequenz": 0.05, "Phase": 0}]
    p["zeitverlaeufe"] = {"spalten": ["t", "u", "y"],
                          "daten": [[0, 1, 2], [50, 50, 50], [20, 20.5, 21]]}
    p["modelle"] = [{"Nr": 1, "Typ": "PTn", "Methode": "Zeit-Prozent",
                     "k_M": 0.8, "n": 3.4, "T_M": 2.2, "Guete": 0.01},
                    {"Nr": 2, "Typ": "PT1TT", "Methode": "Tangente",
                     "k_M": 0.8, "T_T": 1.1, "T_1": 6.5, "Guete": 0.02}]
    p["regler"] = [{"Nr": 1, "Modell_Nr": 1, "Verfahren": "Latzel",
                    "Reglertyp": "PI", "Optionen": "h_m=0.1", "T_A": 0,
                    "k_P": 1.2, "T_N": 4.3}]
    p["kennwerte"] = [{"Regler_Nr": 1, "Simulationsart": "fuehrungssprung",
                       "Toleranzband": 0.05, "h_m": 0.08, "T_aus": 22.5,
                       "u_max": 71.0}]
    q = pd.lesen(pd.schreiben(p, programm="Test", aktion="alles"))
    assert q["stoerungen"][0]["Typ"] == "sinus"
    assert q["stoerungen"][0]["Angriff"] == "ausgang"
    assert q["zeitverlaeufe"]["spalten"] == ["t", "u", "y"]
    assert q["zeitverlaeufe"]["daten"][2] == [20, 20.5, 21]
    assert len(q["modelle"]) == 2
    assert q["modelle"][1]["T_1"] == 6.5
    assert q["regler"][0]["k_P"] == 1.2
    assert q["kennwerte"][0]["T_aus"] == 22.5


def test_fremde_blaetter_bleiben_erhalten():
    """Programm B liest, ergänzt sein Blatt, schreibt – Blatt von A bleibt."""
    p = pd.neu("Fall-C", programm="A")
    p["system"] = dict(BEISPIEL_SYSTEM)
    inhalt_a = pd.schreiben(p, programm="A", aktion="System")
    q = pd.lesen(inhalt_a)
    q["modelle"] = [{"Nr": 1, "Typ": "PTn", "k_M": 1.0, "n": 3, "T_M": 1.0}]
    inhalt_b = pd.schreiben(q, programm="B", aktion="Modelle")
    r = pd.lesen(inhalt_b)
    assert r["system"]["systemname"] == "PT2-Test"     # von A, unangetastet
    assert r["modelle"][0]["n"] == 3                   # von B
    assert [h["programm"] for h in r["meta"]["historie"]] == ["A", "A", "B"]


def test_altformat_import():
    """RKZ-Systemdatei (nur Blatt System) wird beim Lesen überführt."""
    wb = Workbook()
    ws = wb.active
    ws.title = "System"
    ws.append(["Feld", "Wert"])
    labels = dict(pd.SYSTEM_FELDER)
    for key in ("systemname", "a_text", "b_text", "c_text",
                "u_min", "u_max", "k_s", "y_min"):
        ws.append([labels[key], BEISPIEL_SYSTEM[key]])
    buf = io.BytesIO()
    wb.save(buf)

    q = pd.lesen(buf.getvalue())
    assert q["meta"]["formatversion"] == 1
    assert q["meta"]["fallname"] == "PT2-Test"
    assert q["meta"]["angelegt_von"] == "Altformat-Import"
    assert q["system"]["k_s"] == "0.8"


def test_zu_neue_formatversion():
    p = pd.neu("Fall-D", programm="Test")
    p["meta"]["formatversion"] = 2
    inhalt = pd.schreiben(p)
    with pytest.raises(ProjektdateiError, match="Formatversion 2"):
        pd.lesen(inhalt)


def test_keine_agp_datei():
    wb = Workbook()
    wb.active.title = "Irgendwas"
    wb.active.append(["x"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ProjektdateiError, match="AGP-Datei"):
        pd.lesen(buf.getvalue())


def test_dateiname_schema():
    name = pd.dateiname("Fall A/B")
    assert name.startswith("Fall_A_B-20")
    assert name.endswith(".xlsx")
