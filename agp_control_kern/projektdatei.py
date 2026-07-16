"""
projektdatei.py – AGP-Projektdatei, Formatversion 1.

Umsetzung der Spezifikation
RegelkreisZustandsDemo/Spezifikationen/25v07-V1-Spec-AGP-Projektdatei.tex
(beschlossen 2026-07-16, P1–P6):

- EINE wachsende .xlsx-Datei je Schulungsfall; Blätter:
  Meta, System, Stoerungen, Zeitverlaeufe, Modelle, Regler, Kennwerte
- Programme schreiben nur ihre eigenen Blätter und hängen eine Historie-Zeile an
- Zeitbasis in der Datei: Sekunden; Zahlen mit Dezimalpunkt, volle Präzision
- Altformat-Import: Datei ohne "Meta", aber mit "System" in der
  RegelkreisZustandsDemo-Struktur (v0.4/v0.6) wird beim Lesen überführt

Arbeitsmodell: Ein "Projekt" ist ein dict; `lesen()` und `schreiben()` wandeln
zwischen bytes (.xlsx) und dict. Programme ändern nur ihre Einträge im dict.

    projekt = {
      "meta": {"formatversion": 1, "fallname": …, "beschreibung": …,
               "angelegt_am": …, "angelegt_von": …,
               "historie": [{"zeitstempel": …, "programm": …, "aktion": …}, …]},
      "system":        dict  | None,   # Schlüssel/Wert (Felder wie RKZ v0.6)
      "stoerungen":    list  | None,   # Zeilen-dicts, Spalten STOERUNG_SPALTEN
      "zeitverlaeufe": dict  | None,   # {"spalten": […], "daten": [[…], …]}
      "modelle":       list  | None,
      "regler":        list  | None,
      "kennwerte":     list  | None,
    }
"""

from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

FORMATVERSION = 1
MAX_BESCHREIBUNG = 1000


class ProjektdateiError(ValueError):
    """Fehler mit verständlicher Meldung für das Frontend."""


# ─── Blattdefinitionen (Spec V1) ─────────────────────────────────────────────

# System-Blatt: Feldnamen/Labels wie die RKZ-v0.6-Systemdatei → Alt-Import direkt
SYSTEM_FELDER = [
    ("systemname",       "Systemname"),
    ("a_text",           "Systemmatrix A"),
    ("b_text",           "Eingangsvektor b"),
    ("c_text",           "Ausgangsvektor c"),
    ("u_min",            "Stellbereich u_min"),
    ("u_max",            "Stellbereich u_max"),
    ("k_s",              "Systemverstärkung k_S"),
    ("y_min",            "Regelbereich y_min"),
    ("eingabeform",      "Eingabeform"),
    ("zaehler_text",     "G(s) Zählerpolynom"),
    ("nenner_text",      "G(s) Nennerpolynom"),
    ("nullstellen_text", "G(s) Nullstellen"),
    ("pole_text",        "G(s) Pole"),
    ("zaehler_zk_text",  "G(s) Zähler-Zeitkonstanten"),
    ("nenner_zk_text",   "G(s) Nenner-Zeitkonstanten"),
    ("k_faktor",         "G(s) Faktor K"),
    ("normalform",       "Normalform"),
    ("beschreibung",     "Beschreibung"),
]
_SYSTEM_LABEL_ZU_KEY = {label: key for key, label in SYSTEM_FELDER}
SYSTEM_OPTIONAL = {"systemname", "beschreibung", "eingabeform",
                   "zaehler_text", "nenner_text", "nullstellen_text",
                   "pole_text", "zaehler_zk_text", "nenner_zk_text",
                   "k_faktor", "normalform"}

STOERUNG_SPALTEN = ["Nr", "Typ", "Angriff", "Aktiv", "Amplitude", "Startzeit",
                    "Rate", "Frequenz", "Phase", "Seed", "Kommentar"]
MODELL_SPALTEN = ["Nr", "Typ", "Methode", "k_M", "n", "T_M", "T_T", "T_1", "Guete"]
REGLER_SPALTEN = ["Nr", "Modell_Nr", "Verfahren", "Reglertyp", "Optionen",
                  "T_A", "k_P", "T_N", "T_V"]
KENNWERT_SPALTEN = ["Regler_Nr", "Simulationsart", "Toleranzband", "h_m",
                    "T_an", "T_aus", "u_max", "e_bleibend",
                    "Mittelwert_y", "ZweiSigma_y", "Mittelwert_u", "ZweiSigma_u",
                    "Kommentar"]

_TABELLEN = [("Stoerungen", "stoerungen", STOERUNG_SPALTEN),
             ("Modelle", "modelle", MODELL_SPALTEN),
             ("Regler", "regler", REGLER_SPALTEN),
             ("Kennwerte", "kennwerte", KENNWERT_SPALTEN)]

_META_FELDER = [("formatversion", "Formatversion"),
                ("fallname", "Fallname"),
                ("beschreibung", "Beschreibung"),
                ("angelegt_am", "Angelegt_am"),
                ("angelegt_von", "Angelegt_von")]
_META_LABEL_ZU_KEY = {label: key for key, label in _META_FELDER}


# ─── Hilfen ──────────────────────────────────────────────────────────────────

def zeitstempel() -> str:
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")


def dateiname(fallname: str) -> str:
    """Dateinamens-Schema (Beschluss P5): Fallname-Datum_Uhrzeit.xlsx"""
    basis = re.sub(r"\s+", "_", re.sub(r"[^\wäöüÄÖÜß\- ]+", "_",
                                       (fallname or "Fall").strip()))
    return f"{basis}-{zeitstempel()}.xlsx"


def neu(fallname: str, beschreibung: str = "", programm: str = "") -> dict:
    """Leeres Projekt (nur Meta gefüllt)."""
    return {
        "meta": {
            "formatversion": FORMATVERSION,
            "fallname": fallname,
            "beschreibung": beschreibung,
            "angelegt_am": zeitstempel(),
            "angelegt_von": programm,
            "historie": [{"zeitstempel": zeitstempel(), "programm": programm,
                          "aktion": "Projekt angelegt"}],
        },
        "system": None, "stoerungen": None, "zeitverlaeufe": None,
        "modelle": None, "regler": None, "kennwerte": None,
    }


# ─── Schreiben ───────────────────────────────────────────────────────────────

def _kv_blatt(wb: Workbook, titel: str) -> "Worksheet":
    ws = wb.active if (wb.active.max_row == 1 and wb.active.max_column == 1) \
        else wb.create_sheet()
    ws.title = titel
    fett = Font(bold=True)
    ws.append(["Feld", "Wert"])
    ws["A1"].font = fett
    ws["B1"].font = fett
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60
    return ws


def schreiben(projekt: dict, programm: str = "", aktion: str = "") -> bytes:
    """Projekt-dict → .xlsx-Bytes; hängt (wenn programm gesetzt) Historie an."""
    meta = projekt.get("meta") or {}
    if len(str(meta.get("beschreibung", ""))) > MAX_BESCHREIBUNG:
        raise ProjektdateiError(
            f"Beschreibung ist länger als {MAX_BESCHREIBUNG} Zeichen.")
    if programm:
        meta.setdefault("historie", []).append(
            {"zeitstempel": zeitstempel(), "programm": programm,
             "aktion": aktion or "geschrieben"})

    wb = Workbook()
    fett = Font(bold=True)

    # Meta
    ws = _kv_blatt(wb, "Meta")
    for key, label in _META_FELDER:
        ws.append([label, meta.get(key, "")])
    ws.append([])
    ws.append(["Zeitstempel", "Programm", "Aktion"])
    for zelle in ws[ws.max_row]:
        zelle.font = fett
    for h in meta.get("historie", []):
        ws.append([h.get("zeitstempel", ""), h.get("programm", ""),
                   h.get("aktion", "")])

    # System (Schlüssel/Wert)
    if projekt.get("system") is not None:
        ws = _kv_blatt(wb, "System")
        for key, label in SYSTEM_FELDER:
            ws.append([label, projekt["system"].get(key, "")])

    # Tabellen-Blätter
    for blattname, key, spalten in _TABELLEN:
        zeilen = projekt.get(key)
        if zeilen is None:
            continue
        ws = wb.create_sheet(blattname)
        ws.append(spalten)
        for zelle in ws[1]:
            zelle.font = fett
        for zeile in zeilen:
            ws.append([zeile.get(s, "") for s in spalten])

    # Zeitverläufe (Spaltennamen + Datenspalten)
    zv = projekt.get("zeitverlaeufe")
    if zv is not None:
        ws = wb.create_sheet("Zeitverlaeufe")
        spalten, daten = zv["spalten"], zv["daten"]
        if len(spalten) != len(daten):
            raise ProjektdateiError(
                "Zeitverläufe: Spaltennamen und Datenspalten passen nicht zusammen.")
        ws.append(spalten)
        for zelle in ws[1]:
            zelle.font = fett
        n_zeilen = max((len(s) for s in daten), default=0)
        for i in range(n_zeilen):
            ws.append([s[i] if i < len(s) else None for s in daten])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Lesen ───────────────────────────────────────────────────────────────────

def _lese_meta(ws) -> dict:
    meta: dict = {"historie": []}
    historie_modus = False
    for zeile in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        a = zeile[0]
        if a is None or str(a).strip() == "":
            continue
        if str(a) == "Zeitstempel":       # Kopf der Historie-Tabelle
            historie_modus = True
            continue
        if historie_modus:
            meta["historie"].append({"zeitstempel": str(a),
                                     "programm": "" if zeile[1] is None else str(zeile[1]),
                                     "aktion": "" if zeile[2] is None else str(zeile[2])})
        elif str(a) in _META_LABEL_ZU_KEY:
            wert = zeile[1]
            meta[_META_LABEL_ZU_KEY[str(a)]] = "" if wert is None else wert
    try:
        meta["formatversion"] = int(meta.get("formatversion", 0))
    except (TypeError, ValueError):
        raise ProjektdateiError("Meta: Formatversion ist keine Zahl.")
    return meta


def _lese_system(ws) -> dict:
    daten: dict = {}
    for zeile in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        label, wert = zeile[0], zeile[1]
        if label in _SYSTEM_LABEL_ZU_KEY:
            daten[_SYSTEM_LABEL_ZU_KEY[label]] = "" if wert is None else str(wert)
    fehlend = [label for key, label in SYSTEM_FELDER
               if key not in daten and key not in SYSTEM_OPTIONAL]
    if fehlend:
        raise ProjektdateiError(
            "Blatt System: es fehlen die Felder " + ", ".join(fehlend))
    return daten


def _lese_tabelle(ws, spalten: list[str]) -> list[dict]:
    kopf = [str(z.value) if z.value is not None else "" for z in ws[1]]
    zeilen = []
    for werte in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in werte):
            continue
        zeile = {}
        for name, wert in zip(kopf, werte):
            if name in spalten:
                zeile[name] = "" if wert is None else wert
        zeilen.append(zeile)
    return zeilen


def _lese_zeitverlaeufe(ws) -> dict:
    spalten = [str(z.value) for z in ws[1] if z.value is not None]
    daten: list[list] = [[] for _ in spalten]
    for werte in ws.iter_rows(min_row=2, values_only=True):
        for i in range(len(spalten)):
            v = werte[i] if i < len(werte) else None
            if v is not None:
                daten[i].append(v)
    return {"spalten": spalten, "daten": daten}


def lesen(inhalt: bytes) -> dict:
    """Liest eine Projektdatei (oder eine Alt-Systemdatei) zum Projekt-dict."""
    try:
        wb = load_workbook(io.BytesIO(inhalt), data_only=True)
    except Exception:
        raise ProjektdateiError(
            "Datei konnte nicht als Excel-Datei (.xlsx) gelesen werden.")

    projekt = {"meta": None, "system": None, "stoerungen": None,
               "zeitverlaeufe": None, "modelle": None, "regler": None,
               "kennwerte": None}

    if "Meta" in wb.sheetnames:
        projekt["meta"] = _lese_meta(wb["Meta"])
        if projekt["meta"]["formatversion"] > FORMATVERSION:
            raise ProjektdateiError(
                f"Die Datei hat Formatversion {projekt['meta']['formatversion']}; "
                f"dieses Programm kennt nur Version {FORMATVERSION}. "
                "Bitte das Programm aktualisieren.")
    elif "System" in wb.sheetnames:
        # Altformat (RKZ-Systemdatei v0.4/v0.6) → überführen
        system = _lese_system(wb["System"])
        projekt["meta"] = {
            "formatversion": FORMATVERSION,
            "fallname": system.get("systemname") or "Altdatei",
            "beschreibung": system.get("beschreibung", ""),
            "angelegt_am": zeitstempel(),
            "angelegt_von": "Altformat-Import",
            "historie": [{"zeitstempel": zeitstempel(),
                          "programm": "Altformat-Import",
                          "aktion": "Alt-Systemdatei überführt"}],
        }
        projekt["system"] = system
        return projekt
    else:
        raise ProjektdateiError(
            'Weder Blatt "Meta" (Projektdatei) noch Blatt "System" '
            "(Alt-Systemdatei) gefunden – ist das eine AGP-Datei?")

    if "System" in wb.sheetnames:
        projekt["system"] = _lese_system(wb["System"])
    for blattname, key, spalten in _TABELLEN:
        if blattname in wb.sheetnames:
            projekt[key] = _lese_tabelle(wb[blattname], spalten)
    if "Zeitverlaeufe" in wb.sheetnames:
        projekt["zeitverlaeufe"] = _lese_zeitverlaeufe(wb["Zeitverlaeufe"])
    return projekt
